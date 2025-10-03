
import numpy as np
import math as m
from itertools  import permutations
from itertools import product
from functions_BDA import tratin,TMyTEFSC,conttFSC,TMyTEFSF,conttFSF,selector,colsum,vertices,sucessor_sequence, dimensions
import matplotlib.pyplot as plt
import copy
from openpyxl import Workbook
 
axisxc=[]
axisyc=[]
axissc=[]
axisxf=[]
axisyf=[]
axissf=[]
axisx=[]
axisy=[]
axiss=[]
Sucessor_sequence=[]
Centro_list=[]
Dim_algebra=[]
Vertices=[]
Cong=[]
Ent=[]
E=[]
C_list=[]
L_cummulative=[]

# File xlsx
wb = Workbook()
ws = wb.active
ws.title = "Data"
encabezados = ["Configutation", "Release time",'Idle time', "Makespan","Dimension of the algebra","Dimension of the center","Entropy","Loops"]
for col_index, encabezado in enumerate(encabezados, start=1):
    ws.cell(row=1, column=col_index).value = encabezado

#Instance OSSP
matrix1 = np.array([[2,2,2],[2,2,2],[2,2,2]])

numf,numc = matrix1.shape
viaje=8
matrix=matrix1+viaje
matrix[:numf-1] += viaje
sum_rows=[]

for i in range(numc):
    sum1=0
    for j in range(numf):
        sum1=sum1+matrix[j][i]
    sum_rows.append(int(sum1))
    
maxfc,ttotal=selector(matrix)
sumcol=colsum(matrix)   
matrix = tratin(matrix)
nummax = matrix[1]
matrix = matrix[0]
matrixP = []


for i in matrix:
    matrixP.append(list(permutations(i)))    

q=0

for elem in product(*matrixP):
    e=[]
    sum_rows1=sum_rows.copy()
    sum_rows2=sum_rows.copy()
    
    elemod=[list(row) for row in elem]
    elemodc=copy.deepcopy(elemod)

    for index,i in enumerate(elem):
        e.append([])
        for j in i:
            e[index].append(j[1])
    lines = [", ".join(map(str, fila)) for fila in e]


    result = "\n".join(lines)
    
    E.append(e)
    matrixc= TMyTEFSC(elemod,nummax)
    xc,yc=conttFSC(matrixc)
    axisxc.append(xc)
    axisyc.append(yc)
    axissc.append(xc+yc)
    
    
    matrixf= TMyTEFSF(elem,numf,numc)
    xf,yf=conttFSF(matrixf)
    axisxf.append(xf)
    axisyf.append(yf)
    axissf.append(xf+yf)


    if xc+yc<=xf+yf:
        axisx.append(xc)
        axisy.append(yc)
        axiss.append(xc+yc)   
        Vertices=vertices(matrixc)
        Sucessorsequence=sucessor_sequence(matrixc,Vertices)
        sum_rows2.append(axisx[q])
        sum_rows2.append(axisy[q])
        A=matrixc

    else:
        axisx.append(xf)
        axisy.append(yf)
        axiss.append(xf+yf)
        Vertices=vertices(matrixf)  
        Sucessorsequence=sucessor_sequence(matrixf,Vertices)
        sum_rows2.append(axisx[q])
        sum_rows2.append(axisy[q])
        A=matrixf


        
    Sucessor_sequence.append(Sucessorsequence)   
    
    dimension,centro,entropia,loops,Loops=dimensions(sum_rows2,numf, Sucessorsequence,Vertices)
    Dim_algebra.append(dimension)
    Centro_list.append(centro)
    Ent.append(entropia)
    C_list.append(sum_rows2[0]+sum_rows2[1]+sum_rows2[2]+sum_rows2[3]+sum_rows2[4])
    result2 = ' '.join(str(item) for item in Loops)
    nombres=[result, axisx[q],axisy[q], sum_rows2[0]+sum_rows2[1]+sum_rows2[2]+sum_rows2[3]+sum_rows2[4],dimension,centro,entropia,loops]
    for i, nombre in enumerate(nombres):  
        ws.cell(row=q+2, column=i+1).value = nombre         
    


    q+=1

wb.save("BDA_solutions_instance.xlsx")
tiempomin = min(min(axissc),min(axissf))

    
    
    

    
    
    


    
            


            


   
    

    
        
        
        
        
        
        
        
